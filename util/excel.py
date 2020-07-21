# encoding = 'utf-8'
# author = jessica
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font
import time
from project_var.var import *


class ParseExcel(object):

    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.workbook = load_workbook(self.excel_path, data_only=True)  # 加载excel, 当excel中包含公式时，设置data_only=True将读公式的结果而不是读公式
        self.sheet = self.workbook.active  # 获取第一个sheet
        self.font = Font(color=None)
        self.color_dict = {'red': 'FFFF3030', 'green': 'FF008B00'}

    # 获取当前默认sheet的名字
    def get_default_sheet(self):
        return self.sheet.title

    # 设置当前要操作的sheet对象，使用index来获取相应的sheet
    def set_sheet_by_index(self, sheet_index):
        sheet_name = self.workbook.sheetnames[sheet_index]
        self.sheet = self.workbook[sheet_name]  # 一定要写赋值语句，不能直接返回
        return self.sheet

    # 设置当前要操作的sheet对象，使用sheet名称来获取相应的sheet
    def set_sheet_by_name(self, sheet_name):
        self.sheet = self.workbook[sheet_name]  # 一定要写赋值语句，不能直接返回
        return self.sheet

    # 获取默认sheet中最大的行数
    def get_max_row_no(self):
        return self.sheet.max_row

    # 获取默认sheet中最大的列数
    def get_max_col_no(self):
        return self.sheet.max_column

    # 获取默认sheet的最小（起始）行号
    def get_min_row_no(self):
        return self.sheet.min_row

    # 获取默认sheet的最小（起始）列号
    def get_min_col_no(self):
        return self.sheet.min_column

    # 获取默认sheet的所有行对象
    def get_all_rows(self):
        # return list(self.rows) 也可以
        return list(self.sheet.iter_rows())

    # 获取默认sheet的所有列对象
    def get_all_cols(self):
        # return list(self.columns) 也可以
        return list(self.sheet.iter_cols())

    # 从默认sheet中获取某一行，第一行从0开始
    def get_single_row(self, row_no):
        return self.get_all_rows()[row_no]

    # 从默认sheet中获取某一列，第一列从0开始
    def get_single_col(self, col_no):
        return self.get_all_cols()[col_no]

    # 从默认sheet中，通过行号和列号获取指定的单元格，注意行号和列号从1开始
    def get_cell(self, row_no, col_no):
        return self.sheet.cell(row=row_no, column=col_no)

    # 从默认sheet中，通过行号和列号获取指定的单元格中的内容，注意行号和列号从1开始
    def get_cell_content(self, row_no, col_no):
        return self.sheet.cell(row=row_no, column=col_no).value

    # 从默认sheet中，通过行号和列号向指定的单元格中写入指定的内容，注意行号和列号从1开始
    # 调用此方法时，excel不要处于打开状态
    def write_cell_content(self, row_no, col_no, content, font=None):
        self.sheet.cell(row=row_no, column=col_no).value = content
        if font:
            self.font = Font(color=self.color_dict[font])
            self.sheet.cell(row=row_no, column=col_no).font = self.font
        self.workbook.save(self.excel_path)
        return self.sheet.cell(row=row_no, column=col_no).value

    # 从默认sheet中，通过行号和列号向指定的单元格中写入当前日期，注意行号和列号从1开始
    # 调用此方法时，excel不要处于打开状态
    def write_cell_current_time(self, row_no, col_no):
        current_time = time.strftime('%Y-%m-%d %H:%M:%S')
        self.sheet.cell(row=row_no, column=col_no).value = str(current_time)
        self.workbook.save(self.excel_path)
        return self.sheet.cell(row=row_no, column=col_no).value

    # 保存excel文件
    def save_excel_file(self):
        self.workbook.save(self.excel_path)


if __name__ == '__main__':
    # 测试代码
    p = ParseExcel(test_data_excel_path)
    print(u'获取当前默认sheet的名字: ', p.get_default_sheet())
    print(u'设置当前要操作的sheet对象，使用index=1来获取相应的sheet: ', p.set_sheet_by_index(1))
    print(u'获取当前默认sheet的名字: ', p.get_default_sheet())
    print(u'设置当前要操作的sheet对象，使用index=0来获取相应的sheet: ', p.set_sheet_by_index(0))
    print(u'获取当前默认sheet的名字: ', p.get_default_sheet())
    print(u'设置当前要操作的sheet对象，使用sheet名称来获取相应的sheet: ', p.set_sheet_by_name('account'))
    print(u'获取当前默认sheet的名字: ', p.get_default_sheet())
    print(u'获取默认sheet中最大的行数: ', p.get_max_row_no())
    print(u'获取默认sheet中最大的列数: ', p.get_max_col_no())
    print(u'获取默认sheet的最小（起始）行号: ', p.get_min_row_no())
    print(u'获取默认sheet的最小（起始）列号: ', p.get_min_col_no())
    print(u'获取默认sheet的所有行对象: ', p.get_all_rows())
    print(u'获取默认sheet的所有列对象: ', p.get_all_cols())
    print(u'从默认sheet中获取某一行，第一行从0开始: ', p.get_single_row(2))
    print(u'从默认sheet中获取某一列，第一列从0开始: ', p.get_single_col(3))
    print(u'从默认sheet中，通过行号和列号获取指定的单元格(2,2)，注意行号和列号从1开始: ', p.get_cell(2, 2))
    print(u'从默认sheet中，通过行号和列号获取指定的单元格(3,3)中的内容，注意行号和列号从1开始: ', p.get_cell_content(3, 3))
    print(u'从默认sheet中，通过行号和列号向指定的单元格中(3,8)写入指定的内容，注意行号和列号从1开始: ', p.write_cell_content(3, 8, u'成功'))
    print(u'从默认sheet中，通过行号和列号向指定的单元格中(2,7)写入当前日期，注意行号和列号从1开始: ', p.write_cell_current_time(2, 7))
    print(u'保存excel文件: ', p.save_excel_file())
